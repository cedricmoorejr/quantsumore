# -*- coding: utf-8 -*-
#
## ╭────────────────────────────────────────────────────────────────────────────────────────────╮
## │  Library         : doydl's Finance API Client — quantsumore                                 │
## │                                                                                             │
## │                                                                                             │
## │  Description     : `quantsumore` is a comprehensive Python library designed to streamline   │
## │                    the process of accessing and analyzing real-time financial data across   │
## │                    various markets. It provides specialized API clients to fetch data       │
## │                    from multiple financial instruments, including:                          │
## │                      - Cryptocurrencies                                                     │
## │                      - Equities and Stock Markets                                           │
## │                      - Foreign Exchange (Forex)                                             │
## │                      - Treasury Instruments                                                 │
## │                      - Consumer Price Index (CPI) Metrics                                   │
## │                                                                                             │
## │                    The library offers a unified interface for retrieving diverse financial  │
## │                    data, enabling users to perform in-depth financial and technical         │
## │                    analysis. Whether you're developing trading algorithms, conducting       │
## │                    market research, or building financial dashboards, `quantsumore` serves  │
## │                    as a reliable and efficient tool in your data pipeline.                  │
## │                                                                                             │
## │                                                                                             │
## │  Key Features    : - Real-time data retrieval from multiple financial markets               │
## │                    - Support for various financial instruments and metrics                  │
## │                    - Simplified API clients for ease of integration                         │
## │                    - Designed for both personal and non-commercial use                      │
## │                                                                                             │
## │                                                                                             │
## │  Legal Disclaimer: `quantsumore` is an independent Python library and is not affiliated     │
## │                    with any financial institutions or data providers. Likewise, doydl       │
## │                    technologies is not affiliated with, endorsed by, or sponsored by any    │
## │                    government, corporate, or financial institutions. Users should verify    │
## │                    the accuracy of the data obtained and consult professional advice        │
## │                    before making investment decisions.                                      │
## │                                                                                             │
## │                                                                                             │
## │  Copyright       : © 2023–2025 by doydl technologies. All rights reserved.                  │
## │                                                                                             │
## │                                                                                             │
## │  License         : Licensed under the Apache License, Version 2.0 (the "License");          │
## │                    you may not use this file except in compliance with the License.         │
## │                    You may obtain a copy of the License at:                                 │
## │                                                                                             │
## │                        http://www.apache.org/licenses/LICENSE-2.0                           │
## │                                                                                             │
## │                    Unless required by applicable law or agreed to in writing, software      │
## │                    distributed under the License is distributed on an "AS IS" BASIS,        │
## │                    WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or          │
## │                    implied. See the License for the specific language governing             │
## │                    permissions and limitations under the License.                           │
## ╰────────────────────────────────────────────────────────────────────────────────────────────╯
#

"""
_frameworks: Statement Layout Registry and Excel Export Engine for Financials
══════════════════════════════════════════════════════════════════════════════

Module Purpose
────────────────────────────────────────────────────
`_frameworks` acts as the **central schema registry** for the formatting,
hierarchy, and display logic of all core financial statements in the
Quantsumore ecosystem. It also provides a robust `WriteExcel` engine
to **export financial dataframes into styled, human-readable Excel sheets**,
with auto-indentation, subtotal formatting, column width control, and
professional typographic conventions.

This module is used internally by all higher-level analysis and reporting
functions that require exporting company fundamentals, summary tables, or
report packages in `.xlsx` format.

Core Components
────────────────────────────────────────────────────

1. **`_STATEMENT_LAYOUTS` Registry**
   • A single-source dictionary specifying the hierarchical structure and display
     logic for the three fundamental statements: **Income Statement**, **Balance Sheet**,
     and **Cash Flow Statement**.
   • Each statement includes:
     • `indentation_levels`: Dict of account labels to indentation depth (for visual tree structure)
     • `parent_accounts`: Account labels serving as top-level categories (rendered bold)
     • `subtotal_total_accounts`: Account labels for subtotals/totals (rendered bold with special borders)

2. **`WriteExcel` Class**
   • The main interface for writing styled financial statements to Excel.
   • Handles sheet creation/replacement, data writing, cell styling, alignment, and column width fitting.
   • Applies a consistent style guide (font, size, bolding, fills, and border logic) so exports always look professional.

Key Use Cases
────────────────────────────────────────────────────
	• Batch or on-demand export of cleaned DataFrames (Income, Balance, Cash Flow) to `.xlsx` files.
	• Rendering of hierarchical statements with account-level indentation and subtotal logic for readability.
	• Automated workflows that require visually consistent financial exports for compliance, reporting, or sharing.
	• Direct use as a backend by dashboards or UI apps that let users download raw or formatted financials.

System Architecture & Styling Logic
────────────────────────────────────────────────────
	• Uses OpenPyXL for all Excel I/O and cell styling.
	• Column widths are **auto-fit** to data and headers for optimal reading, never requiring manual adjustment.
	• Account names are indented according to their logical depth in the hierarchy (`indentation_levels`).
	• Parent accounts and all subtotal/total rows are **bolded** for rapid identification.
	• Subtotal rows receive a thin bottom border; total rows receive a double border for clear demarcation.
	• Numeric cells are right-aligned and formatted with two decimal places.
	• Sheet-level styles (named styles) are dynamically registered if missing, so workbook is always standards-compliant.

Design Features
────────────────────────────────────────────────────
	• **Centralized schema**: All statement formatting, grouping, and subtotal logic is defined in one place for maintainability.
	• **Idempotent export**: Existing sheets are replaced by default (or auto-suffixed if overwrite is off), so repeated exports never break.
	• **Separation of layout from logic**: Data structure (`_STATEMENT_LAYOUTS`) and export logic (`WriteExcel`) are fully decoupled.
	• **Type- and error-safety**: Non-DataFrame inputs or bad filenames are handled with descriptive exceptions.

Statement Layouts
────────────────────────────────────────────────────
	• "Income Statement" — supports nested expenses, non-recurring items, and correct net income positioning
	• "Balance Sheet" — handles multi-level assets/liabilities, equity, and all common subtotals/totals
	• "Cash Flow Statement" — supports all major groupings (operating, investing, financing) and subtotal/total logic

Available Classes & Exports
────────────────────────────────────────────────────
	• WriteExcel — context-managed Excel workbook writer
	• _STATEMENT_LAYOUTS — dict for layout, indentation, and subtotal logic

Implementation Notes
────────────────────────────────────────────────────
	• Sheet detection is automated by scanning index labels for key accounts
	• All styles (font, bold, alignment, fill, borders) are registered as named styles at workbook level
	• Can be safely used in Jupyter, scripts, or backend pipelines with large data
"""
import os
from datetime import datetime
import random

# ────────── Project-specific imports (directly from this project's source code) ─────────────────────────────
from ..exceptions import WorkbookSaveError


__all__ = ['WriteExcel']



# ━━━━━━━━━━━━━━ Core Module Implementation ━━━━━━━━━━━━━━━━━━━━━━━━━━
# This segment delineates the functional backbone of the module.
# It comprises the abstractions and behaviors essential for runtime
# execution—if applicable—encapsulated in class and function constructs.
# In minimal implementations, this may simply define constants, metadata,
# or serve as an interface placeholder.

_STATEMENT_LAYOUTS = {
    "Income Statement": {
        "indentation_levels": {
            "Total Revenue":0, "Cost of Revenue":1, "Gross Profit":0, "Operating Expenses":0,
            "Research and Development":1, "Sales, General and Admin.":1, "Non-Recurring Items":1,
            "Other Operating Items":1, "Operating Income":0, "Add\\'l income/expense items":1,
            "Earnings Before Interest and Tax":0, "Interest Expense":1, "Earnings Before Tax":0,
            "Income Tax":1, "Minority Interest":1, "Equity Earnings/Loss Unconsolidated Subsidiary":1,
            "Net Income-Cont. Operations":0, "Net Income":0, "Net Income Applicable to Common Shareholders":0
        },
        "parent_accounts": ["Total Revenue", "Operating Expenses"],
        "subtotal_total_accounts": [
            "Gross Profit","Operating Income","Earnings Before Interest and Tax",
            "Earnings Before Tax","Net Income-Cont. Operations","Net Income",
            "Net Income Applicable to Common Shareholders"
        ]
    },
    "Balance Sheet": {
        "indentation_levels": {
            "Current Assets":0, "Cash and Cash Equivalents":1, "Short-Term Investments":1,
            "Net Receivables":1, "Inventory":1, "Other Current Assets":1, "Total Current Assets":0,
            "Long-Term Assets":0, "Long-Term Investments":1, "Fixed Assets":1, "Goodwill":1,
            "Intangible Assets":1, "Other Assets":1, "Deferred Asset Charges":1, "Total Assets":0,
            "Current Liabilities":0, "Accounts Payable":1, "Short-Term Debt / Current Portion of Long-Term Debt":1,
            "Other Current Liabilities":1, "Total Current Liabilities":0, "Long-Term Debt":0, "Other Liabilities":0,
            "Deferred Liability Charges":1, "Misc. Stocks":1, "Minority Interest":1, "Total Liabilities":0,
            "Stock Holders Equity":0, "Common Stocks":1, "Capital Surplus":1, "Retained Earnings":1,
            "Treasury Stock":1, "Other Equity":1, "Total Equity":0, "Total Liabilities & Equity":0
        },
        "parent_accounts": ["Current Assets","Long-Term Assets","Current Liabilities","Stock Holders Equity"],
        "subtotal_total_accounts": [
            "Total Current Assets","Total Assets","Total Current Liabilities",
            "Total Liabilities","Total Equity","Total Liabilities & Equity"
        ]
    },
    "Cash Flow Statement": {
        "indentation_levels": {
            "Net Income":0, "Cash Flows-Operating Activities":0, "Depreciation":1,
            "Net Income Adjustments":1, "Changes in Operating Activities":0, "Accounts Receivable":1,
            "Changes in Inventories":1, "Other Operating Activities":1, "Liabilities":1,
            "Net Cash Flow-Operating":0, "Cash Flows-Investing Activities":0, "Capital Expenditures":1,
            "Investments":1, "Other Investing Activities":1, "Net Cash Flows-Investing":0,
            "Cash Flows-Financing Activities":0, "Sale and Purchase of Stock":1, "Net Borrowings":1,
            "Other Financing Activities":1, "Net Cash Flows-Financing":0, "Effect of Exchange Rate":0,
            "Net Cash Flow":0
        },
        "parent_accounts": [
            "Cash Flows-Operating Activities", "Changes in Operating Activities",
            "Cash Flows-Investing Activities", "Cash Flows-Financing Activities"
        ],
        "subtotal_total_accounts": [
            "Net Cash Flow-Operating","Net Cash Flows-Investing",
            "Net Cash Flows-Financing","Net Cash Flow"
        ]
    }
}



"""
Provides a set of functionalities for writing financial statments to an Excel file, adjusting column widths
automatically based on the data content, and applying various styles to the data to enhance readability and presentation.
"""
# Notes:
# -----
# Regarding the __create_or_replace_sheet method.
# Creates a new sheet in the workbook. If a sheet with the same name exists, it can either
# overwrite it (delete and create new) or create a new sheet with a unique name.
# 
# Parameters:
# - sheet_name (str): Name of the sheet to create.
# - overwrite (bool): If True, deletes the existing sheet with the same name.
#                     If False, a new sheet with an appended suffix (like '_1') is created.
class WriteExcel:
    def __init__(self, path=None):
        self.path = path
        if path and os.path.exists(path):
            from openpyxl import load_workbook; self.wb = load_workbook(path)
        else:
            from openpyxl import Workbook; self.wb = Workbook()
            if 'Sheet' in self.wb.sheetnames: del self.wb['Sheet']
        self.base_width, self.min_width = 1.0, 10

    def __enter__(self): return self
    def __exit__(self, *a): self.wb.close()

    def __which_statement(self, financial_statement):
        a = list(df.index.values)
        if "Research and Development" in a: return "Income Statement"
        if "Goodwill" in a: return "Balance Sheet"
        if "Depreciation" in a: return "Cash Flow Statement"
        return None

    def __coerce_numeric_min(self, n):
        try:
            v = float(n) if '.' in str(n) else int(n)
            return max(int(v) if float(v).is_integer() else v, 1)
        except: return 11
           
    def __autofit_column_width(self, ws, df):
        from openpyxl.utils import get_column_letter
        for i, col in enumerate(df.columns, 1):
            maxlen = max(len(str(cell)) for cell in df[col])
            ws.column_dimensions[get_column_letter(i)].width = max(max(len(col), maxlen)*self.base_width, self.min_width)
                
    def __apply_styles(self, sheet_name, font_style="Calibri Light", font_size=10):
        from openpyxl.styles import NamedStyle, Font
        size = self.__coerce_numeric_min(size)
        s = lambda n, **k: NamedStyle(name=n, font=Font(**k, name=font, size=size), number_format="_(* #,##0.00_);_(* (#,##0.00);_(* \"-\"??_);_(@_)")
        styles = [
            NamedStyle(name=f"data_font_style_{sheet}", font=Font(bold=False, name=font, size=size), number_format="@"),
            NamedStyle(name=f"header_font_style_{sheet}", font=Font(bold=True, name=font, size=size), number_format="@"),
            s(f"number_format_style_{sheet}", bold=False),
            s(f"bold_font_style_{sheet}", bold=True),
            s(f"bold_number_format_style_{sheet}", bold=True)
        ]
        for style in styles:
            if style.name not in self.wb.named_styles: self.wb.add_named_style(style)
                
    def __create_or_replace_sheet(self, sheet_name, overwrite=True):
        if name in self.wb.sheetnames:
            if overwrite: del self.wb[name]
            else:
                s, new = 1, f"{name}_1"
                while new in self.wb.sheetnames: s += 1; new = f"{name}_{s}"
                name = new
        return self.wb.create_sheet(title=name)
        
    def write_statement(self, df, reporting_structure=_STATEMENT_LAYOUTS):
        from openpyxl.styles import Alignment, PatternFill, Border, Side
        sheet = self.__which_statement(df)
        ws = self.__create_or_replace_sheet(sheet, overwrite=True)
        self.__apply_styles(sheet)
        ra = Alignment(horizontal="right")
        fill = PatternFill(start_color='DADADA', end_color='DADADA', fill_type='solid')
        ws.cell(row=1, column=1, value=df.index.name or "Ending:").style = f"header_font_style_{sheet}"
        ws.cell(row=1, column=1).fill = fill
        inds = reporting_structure[sheet]["indentation_levels"]
        parents = reporting_structure[sheet]["parent_accounts"]
        totals = reporting_structure[sheet]["subtotal_total_accounts"]
        for ci, col in enumerate(df.columns, 2):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.style, cell.alignment = f"header_font_style_{sheet}", ra
        for ri, (idx, row) in enumerate(df.iterrows(), 2):
            ind = inds.get(idx.strip(), 0)
            is_bold = idx in parents or idx in totals
            rn = f"{'    '*ind}{idx}"
            ncell = ws.cell(row=ri, column=1, value=rn)
            ncell.style = f"bold_font_style_{sheet}" if is_bold else f"data_font_style_{sheet}"
            for ci, val in enumerate(row, 2):
                c = ws.cell(row=ri, column=ci, value=val)
                c.style = f"bold_number_format_style_{sheet}" if is_bold else f"number_format_style_{sheet}"
                c.alignment = ra
                if idx in totals[:-1]: c.border = Border(bottom=Side(style="thin"))
                elif idx in totals[-1]: c.border = Border(bottom=Side(style="double"))
        self.__autofit_column_width(ws, df)

    def save(self, filename=None, overwrite=True):
        try:
            if not self.wb.sheetnames: self.wb.create_sheet('Sheet1')
            filename = filename or f'output_{datetime.now().strftime("%Y-%m-%d %H_%M_%S")}_{random.randint(1000,5000)}.xlsx'
            if not filename.endswith(".xlsx"): filename += ".xlsx"
            if not overwrite and os.path.exists(filename): raise FileExistsError(f"{filename} exists and 'overwrite'=False.")
            self.wb.save(filename)
            print(f"File '{filename}' has been saved successfully.")
            return True
        except Exception as e:
            raise WorkbookSaveError(f"Failed to save workbook to '{filename}': {e}")

    def close(self): self.wb.close()

def __dir__():
    return __all__
